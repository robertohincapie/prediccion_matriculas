"""Ingesta reproducible de archivos de matrículas.

Responsabilidades de esta etapa:

1. descargar el archivo desde una fuente externa;
2. conservar exactamente los bytes recibidos;
3. calcular SHA-256;
4. verificar que el archivo pueda abrirse como libro Excel;
5. registrar un manifest.json;
6. evitar sobrescrituras silenciosas.

No se realiza limpieza ni transformación de datos.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import unquote, urlparse

import requests
from openpyxl import load_workbook

from matriculas.config import DEFAULT_RAW_DIR, DEFAULT_SOURCE_URL

CHUNK_SIZE = 1024 * 1024
REQUEST_TIMEOUT_SECONDS = 60


class IngestionError(RuntimeError):
    """Error controlado durante la ingesta."""


def filename_from_url(url: str) -> str:
    """Obtiene el nombre del archivo a partir de la URL."""
    path = unquote(urlparse(url).path)
    name = Path(path).name

    if not name:
        raise IngestionError("La URL no contiene un nombre de archivo.")

    return name


def sha256_file(path: Path) -> str:
    """Calcula el SHA-256 de un archivo."""
    digest = hashlib.sha256()

    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(CHUNK_SIZE), b""):
            digest.update(chunk)

    return digest.hexdigest()


def download_to_temp(url: str, temp_dir: Path) -> tuple[Path, dict[str, str]]:
    """Descarga la URL a un archivo temporal sin modificar su contenido."""
    filename = filename_from_url(url)
    temp_path = temp_dir / filename

    try:
        with requests.get(
            url,
            stream=True,
            timeout=REQUEST_TIMEOUT_SECONDS,
            allow_redirects=True,
        ) as response:
            response.raise_for_status()

            headers = {
                "content_type": response.headers.get("Content-Type", ""),
                "content_length": response.headers.get("Content-Length", ""),
                "last_modified": response.headers.get("Last-Modified", ""),
                "etag": response.headers.get("ETag", ""),
                "final_url": response.url,
            }

            with temp_path.open("wb") as output:
                for chunk in response.iter_content(chunk_size=CHUNK_SIZE):
                    if chunk:
                        output.write(chunk)

    except requests.RequestException as exc:
        raise IngestionError(f"No fue posible descargar {url}: {exc}") from exc

    if temp_path.stat().st_size == 0:
        raise IngestionError("El servidor devolvió un archivo vacío.")

    return temp_path, headers


def inspect_excel(path: Path) -> list[str]:
    """Comprueba que el archivo pueda abrirse como Excel y devuelve sus hojas.

    Esta función NO interpreta todavía las columnas ni modifica el workbook.
    """
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet_names = list(workbook.sheetnames)
        workbook.close()
        return sheet_names
    except Exception as exc:
        raise IngestionError(
            f"El archivo descargado no pudo abrirse como libro Excel válido: {exc}"
        ) from exc


def build_manifest(
    *,
    source_url: str,
    stored_path: Path,
    sha256: str,
    sheet_names: list[str],
    response_headers: dict[str, str],
) -> dict:
    """Construye la metadata de trazabilidad de una ingesta."""
    return {
        "source_url": source_url,
        "final_url": response_headers.get("final_url", source_url),
        "filename": stored_path.name,
        "ingested_at_utc": datetime.now(timezone.utc).isoformat(),
        "size_bytes": stored_path.stat().st_size,
        "sha256": sha256,
        "server_metadata": {
            "content_type": response_headers.get("content_type", ""),
            "content_length": response_headers.get("content_length", ""),
            "last_modified": response_headers.get("last_modified", ""),
            "etag": response_headers.get("etag", ""),
        },
        "excel": {
            "sheet_names": sheet_names,
        },
        "status": "raw_ingested",
    }


def ingest(url: str, raw_dir: Path) -> Path:
    """Ingiere un archivo y devuelve la carpeta donde quedó almacenado."""
    filename = filename_from_url(url)

    if Path(filename).suffix.lower() != ".xlsx":
        raise IngestionError(
            f"Esta versión del laboratorio espera un archivo .xlsx; recibido: {filename}"
        )

    dataset_name = Path(filename).stem
    destination_dir = raw_dir / dataset_name
    destination_file = destination_dir / filename
    manifest_path = destination_dir / "manifest.json"

    raw_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="matriculas_ingest_") as tmp:
        temp_dir = Path(tmp)
        temp_file, headers = download_to_temp(url, temp_dir)

        new_hash = sha256_file(temp_file)
        sheet_names = inspect_excel(temp_file)

        if destination_file.exists():
            existing_hash = sha256_file(destination_file)

            if existing_hash == new_hash:
                print("El dataset ya había sido ingerido y su SHA-256 coincide.")
                print(f"Archivo: {destination_file}")
                print(f"SHA-256: {existing_hash}")
                return destination_dir

            raise IngestionError(
                "Ya existe un archivo con el mismo nombre pero diferente contenido. "
                "La ingesta se detuvo para no sobrescribir datos raw."
            )

        destination_dir.mkdir(parents=True, exist_ok=True)
        shutil.move(str(temp_file), destination_file)

        manifest = build_manifest(
            source_url=url,
            stored_path=destination_file,
            sha256=new_hash,
            sheet_names=sheet_names,
            response_headers=headers,
        )

        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    print("Ingesta completada.")
    print(f"Archivo: {destination_file}")
    print(f"Manifest: {manifest_path}")
    print(f"SHA-256: {new_hash}")
    print(f"Hojas Excel: {', '.join(sheet_names) if sheet_names else '(ninguna)'}")

    return destination_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Descarga reproduciblemente un reporte de matrículas."
    )
    parser.add_argument(
        "--url",
        default=DEFAULT_SOURCE_URL,
        help="URL del archivo Excel.",
    )
    parser.add_argument(
        "--raw-dir",
        type=Path,
        default=DEFAULT_RAW_DIR,
        help="Directorio donde conservar los datos raw.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    try:
        ingest(args.url, args.raw_dir)
    except IngestionError as exc:
        raise SystemExit(f"ERROR DE INGESTA: {exc}") from exc


if __name__ == "__main__":
    main()
